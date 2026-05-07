"""
MLB Betting Model — build_model.py
Generates MLB_Betting_Model.xlsx using openpyxl.
TO UPDATE: All editable data is in the CONFIGURATION section below.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, PatternFill, Font, Border, Side
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURATION — edit here to update the model
# ─────────────────────────────────────────────

TEAMS = [
    "Arizona Diamondbacks", "Atlanta Braves", "Baltimore Orioles", "Boston Red Sox",
    "Chicago Cubs", "Chicago White Sox", "Cincinnati Reds", "Cleveland Guardians",
    "Colorado Rockies", "Detroit Tigers", "Houston Astros", "Kansas City Royals",
    "Los Angeles Angels", "Los Angeles Dodgers", "Miami Marlins", "Milwaukee Brewers",
    "Minnesota Twins", "New York Mets", "New York Yankees", "Oakland Athletics",
    "Philadelphia Phillies", "Pittsburgh Pirates", "San Diego Padres", "San Francisco Giants",
    "Seattle Mariners", "St. Louis Cardinals", "Tampa Bay Rays", "Texas Rangers",
    "Toronto Blue Jays", "Washington Nationals",
]

PARK_FACTORS = {
    "Arizona Diamondbacks":  ("Chase Field", 99),
    "Atlanta Braves":        ("Truist Park", 102),
    "Baltimore Orioles":     ("Oriole Park", 98),
    "Boston Red Sox":        ("Fenway Park", 106),
    "Chicago Cubs":          ("Wrigley Field", 103),
    "Chicago White Sox":     ("Guaranteed Rate Field", 97),
    "Cincinnati Reds":       ("Great American Ball Park", 105),
    "Cleveland Guardians":   ("Progressive Field", 97),
    "Colorado Rockies":      ("Coors Field", 115),
    "Detroit Tigers":        ("Comerica Park", 99),
    "Houston Astros":        ("Minute Maid Park", 101),
    "Kansas City Royals":    ("Kauffman Stadium", 98),
    "Los Angeles Angels":    ("Angel Stadium", 98),
    "Los Angeles Dodgers":   ("Dodger Stadium", 101),
    "Miami Marlins":         ("loanDepot park", 97),
    "Milwaukee Brewers":     ("American Family Field", 100),
    "Minnesota Twins":       ("Target Field", 99),
    "New York Mets":         ("Citi Field", 99),
    "New York Yankees":      ("Yankee Stadium", 102),
    "Oakland Athletics":     ("Oakland Coliseum", 96),
    "Philadelphia Phillies": ("Citizens Bank Park", 103),
    "Pittsburgh Pirates":    ("PNC Park", 98),
    "San Diego Padres":      ("Petco Park", 97),
    "San Francisco Giants":  ("Oracle Park", 99),
    "Seattle Mariners":      ("T-Mobile Park", 98),
    "St. Louis Cardinals":   ("Busch Stadium", 100),
    "Tampa Bay Rays":        ("Tropicana Field", 97),
    "Texas Rangers":         ("Globe Life Field", 104),
    "Toronto Blue Jays":     ("Rogers Centre", 100),
    "Washington Nationals":  ("Nationals Park", 97),
}

LEAGUE_AVERAGES = {
    "FIP": 4.20, "xFIP": 4.15, "SIERA": 4.10,
    "K9": 8.9, "BB9": 3.2,
    "wRC_plus": 100, "wOBA": 0.315, "BarrelRate": 8.5,
    "BullpenERA": 4.30,
}

# ─────────────────────────────────────────────
# COLOR CONSTANTS
# ─────────────────────────────────────────────

NAVY        = "1B2A4A"
WHITE       = "FFFFFF"
GREEN       = "2E7D32"
LIGHT_YELLOW = "FFFDE7"
LIGHT_GRAY  = "F5F5F5"
GOLD        = "FFD700"
LIGHT_GREEN = "C8E6C9"
LIGHT_RED   = "FFCDD2"
LIGHT_ORANGE = "FFE0B2"
GRAY_BLUE   = "607D8B"
BLUE        = "0288D1"
DARK_GOLD   = "F9A825"

# ─────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────

def fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header(ws, cell_ref, text, merge_to=None, size=16):
    """Full-width navy header with white bold text."""
    cell = ws[cell_ref]
    cell.value = text
    cell.fill = fill(NAVY)
    cell.font = font(bold=True, color=WHITE, size=size)
    cell.alignment = center()
    if merge_to:
        ws.merge_cells(f"{cell_ref}:{merge_to}")

def apply_section(ws, cell_ref, text, merge_to=None):
    """Section sub-header — navy bg, white bold, size 11."""
    cell = ws[cell_ref]
    cell.value = text
    cell.fill = fill(NAVY)
    cell.font = font(bold=True, color=WHITE, size=11)
    cell.alignment = left()
    if merge_to:
        ws.merge_cells(f"{cell_ref}:{merge_to}")

def input_cell(ws, cell_ref, value=None):
    """Yellow input cell."""
    c = ws[cell_ref]
    if value is not None:
        c.value = value
    c.fill = fill(LIGHT_YELLOW)
    c.font = font(color="000000", size=11)
    c.alignment = left()
    c.border = thin_border()

def calc_cell(ws, cell_ref, formula):
    """Light-gray calculated cell."""
    c = ws[cell_ref]
    c.value = formula
    c.fill = fill(LIGHT_GRAY)
    c.font = font(color="000000", size=11)
    c.alignment = center()
    c.border = thin_border()

def label_cell(ws, cell_ref, text):
    c = ws[cell_ref]
    c.value = text
    c.font = font(color="000000", size=11)
    c.alignment = left()

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ─────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────

wb = Workbook()

# ══════════════════════════════════════════════
# SHEET: Lists  (hidden — team name source)
# ══════════════════════════════════════════════

ws_lists = wb.active
ws_lists.title = "Lists"
ws_lists.sheet_state = "hidden"
ws_lists.tab_color = "AAAAAA"

for i, team in enumerate(TEAMS, start=1):
    ws_lists.cell(row=i, column=1, value=team)

# ══════════════════════════════════════════════
# SHEET 1 — Game Input
# ══════════════════════════════════════════════

ws_gi = wb.create_sheet("Game Input")
ws_gi.tab_color = NAVY
ws_gi.freeze_panes = "A2"

set_col_width(ws_gi, "A", 25)
set_col_width(ws_gi, "B", 20)
set_col_width(ws_gi, "C", 20)
set_col_width(ws_gi, "D", 20)
set_col_width(ws_gi, "E", 20)

# Row 1 — main header
apply_header(ws_gi, "A1", "MLB BETTING MODEL — GAME INPUT", merge_to="E1", size=16)
ws_gi.row_dimensions[1].height = 30

# Row 2 — blank
ws_gi.row_dimensions[2].height = 8

# Row 3 — column sub-headers
label_cell(ws_gi, "A3", "MATCHUP")
ws_gi["A3"].font = font(bold=True, color="000000", size=11)
for col, text in [("B3", "HOME"), ("C3", "AWAY")]:
    ws_gi[col].value = text
    ws_gi[col].fill = fill(NAVY)
    ws_gi[col].font = font(bold=True, color=WHITE, size=11)
    ws_gi[col].alignment = center()

# Row 4 — Team dropdowns
label_cell(ws_gi, "A4", "Team")
input_cell(ws_gi, "B4")
input_cell(ws_gi, "C4")

dv_team = DataValidation(
    type="list",
    formula1="Lists!$A$1:$A$30",
    showDropDown=False,
    allow_blank=True,
)
dv_team.sqref = "B4 C4"
ws_gi.add_data_validation(dv_team)

# Row 5 — Starting Pitcher
label_cell(ws_gi, "A5", "Starting Pitcher")
input_cell(ws_gi, "B5")
input_cell(ws_gi, "C5")

# Row 6 — blank
ws_gi.row_dimensions[6].height = 8

# Row 7 — PITCHER STATS header
apply_section(ws_gi, "A7", "PITCHER STATS", merge_to="C7")

# Rows 8–17 — pitcher stat inputs
pitcher_rows = [
    (8,  "ERA"),
    (9,  "FIP"),
    (10, "xFIP"),
    (11, "SIERA"),
    (12, "K/9"),
    (13, "BB/9"),
    (14, "BABIP"),
    (15, "HR/FB%"),
    (16, "GB%"),
    (17, "Recent Form (Last 5 ERA)"),
]
for row, label in pitcher_rows:
    label_cell(ws_gi, f"A{row}", label)
    input_cell(ws_gi, f"B{row}")
    input_cell(ws_gi, f"C{row}")

# Row 18 — blank
ws_gi.row_dimensions[18].height = 8

# Row 19 — TEAM OFFENSE header
apply_section(ws_gi, "A19", "TEAM OFFENSE", merge_to="C19")

offense_rows = [
    (20, "wOBA"),
    (21, "wRC+"),
    (22, "OPS"),
    (23, "Team BABIP"),
    (24, "Barrel Rate %"),
]
for row, label in offense_rows:
    label_cell(ws_gi, f"A{row}", label)
    input_cell(ws_gi, f"B{row}")
    input_cell(ws_gi, f"C{row}")

# Row 25 — blank
ws_gi.row_dimensions[25].height = 8

# Row 26 — BULLPEN header
apply_section(ws_gi, "A26", "BULLPEN", merge_to="C26")

bullpen_rows = [
    (27, "Bullpen ERA (Last 14 Days)"),
    (28, "Bullpen WHIP"),
]
for row, label in bullpen_rows:
    label_cell(ws_gi, f"A{row}", label)
    input_cell(ws_gi, f"B{row}")
    input_cell(ws_gi, f"C{row}")

# Row 29 — blank
ws_gi.row_dimensions[29].height = 8

# Row 30 — CONTEXT header
apply_section(ws_gi, "A30", "CONTEXT", merge_to="C30")

# Row 31 — Park Factor (auto-calc, gray)
label_cell(ws_gi, "A31", "Park Factor")
calc_cell(ws_gi, "B31", "=IFERROR(VLOOKUP(B4,'Model Engine'!$A$2:$C$31,3,FALSE),100)")
calc_cell(ws_gi, "C31", "=IFERROR(VLOOKUP(C4,'Model Engine'!$A$2:$C$31,3,FALSE),100)")

# Row 32 — Wind Speed
label_cell(ws_gi, "A32", "Wind Speed (mph)")
input_cell(ws_gi, "B32")

# Row 33 — Wind Direction dropdown
label_cell(ws_gi, "A33", "Wind Direction")
input_cell(ws_gi, "B33")
dv_wind = DataValidation(
    type="list",
    formula1='"In,Out,Cross"',
    showDropDown=False,
    allow_blank=True,
)
dv_wind.sqref = "B33"
ws_gi.add_data_validation(dv_wind)

# Row 34 — Temperature
label_cell(ws_gi, "A34", "Temperature (°F)")
input_cell(ws_gi, "B34")

# Row 35 — Days of Rest (Home)
label_cell(ws_gi, "A35", "Days of Rest (Home)")
input_cell(ws_gi, "B35")

# Row 36 — Days of Rest (Away)
label_cell(ws_gi, "A36", "Days of Rest (Away)")
input_cell(ws_gi, "C36")

# Row 37 — blank
ws_gi.row_dimensions[37].height = 8

# Row 38 — VEGAS LINES header
apply_section(ws_gi, "A38", "VEGAS LINES", merge_to="C38")

# Row 39 — Moneyline Odds
label_cell(ws_gi, "A39", "Moneyline Odds")
input_cell(ws_gi, "B39")
input_cell(ws_gi, "C39")

# Row 40 — blank
ws_gi.row_dimensions[40].height = 8

# Row 41 — Navigation hint
ws_gi["A41"].value = "→ Go to Results Dashboard sheet to see model output"
ws_gi["A41"].font = Font(italic=True, color=GREEN, size=11)
ws_gi["A41"].alignment = left()
ws_gi.merge_cells("A41:E41")

# ══════════════════════════════════════════════
# SHEET 2 — Model Engine
# ══════════════════════════════════════════════

ws_me = wb.create_sheet("Model Engine")
ws_me.tab_color = GRAY_BLUE
ws_me.freeze_panes = "A2"

set_col_width(ws_me, "A", 30)
set_col_width(ws_me, "B", 20)
set_col_width(ws_me, "C", 20)

# ── Section A: Park Factor Lookup Table ──────

# Row 1 — headers
for col, text in [("A1", "Team"), ("B1", "Park Name"), ("C1", "Park Factor")]:
    ws_me[col].value = text
    ws_me[col].fill = fill(NAVY)
    ws_me[col].font = font(bold=True, color=WHITE, size=11)
    ws_me[col].alignment = center()
    ws_me[col].border = thin_border()

# Rows 2–31 — all 30 teams (sorted alphabetically = same as TEAMS list, which is already alphabetical)
sorted_teams = sorted(TEAMS)
for i, team in enumerate(sorted_teams, start=2):
    park_name, park_factor = PARK_FACTORS[team]
    ws_me.cell(row=i, column=1, value=team).border = thin_border()
    ws_me.cell(row=i, column=1).font = font(color="000000", size=11)
    ws_me.cell(row=i, column=2, value=park_name).border = thin_border()
    ws_me.cell(row=i, column=2).font = font(color="000000", size=11)
    ws_me.cell(row=i, column=3, value=park_factor).border = thin_border()
    ws_me.cell(row=i, column=3).font = font(color="000000", size=11)
    ws_me.cell(row=i, column=3).alignment = center()
    # Alternate row shading
    if i % 2 == 0:
        for c in range(1, 4):
            ws_me.cell(row=i, column=c).fill = fill("EEF2F7")

# ── Section B: Model Calculations ────────────

# Row 34 header
apply_section(ws_me, "A34", "MODEL CALCULATIONS", merge_to="C34")
ws_me.row_dimensions[34].height = 20

def eng_label(row, text):
    c = ws_me.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color="000000", size=11)
    c.alignment = left()
    c.fill = fill("E8EDF3")
    c.border = thin_border()

def eng_formula(row, formula):
    c = ws_me.cell(row=row, column=2, value=formula)
    c.fill = fill(LIGHT_GRAY)
    c.font = Font(color="000000", size=11)
    c.alignment = center()
    c.border = thin_border()

# ── Pitcher Scores ────────────────────────────

# Row 35: Home Pitcher Score label
eng_label(35, "Home Pitcher Score")

# Row 36: Home Pitcher Score formula
eng_formula(36, (
    "=MAX(0,MIN(100,"
    "(((4.20-'Game Input'!B9)/4.20)*30)"
    "+(((4.15-'Game Input'!B10)/4.15)*25)"
    "+(((4.10-'Game Input'!B11)/4.10)*20)"
    "+(((('Game Input'!B12)-8.9)/8.9)*10)"
    "+(((3.2-'Game Input'!B13)/3.2)*10)"
    "+(((4.20-'Game Input'!B17)/4.20)*5)"
    "+50))"
))

# Row 37: Away Pitcher Score label
eng_label(37, "Away Pitcher Score")

# Row 38: Away Pitcher Score formula
eng_formula(38, (
    "=MAX(0,MIN(100,"
    "(((4.20-'Game Input'!C9)/4.20)*30)"
    "+(((4.15-'Game Input'!C10)/4.15)*25)"
    "+(((4.10-'Game Input'!C11)/4.10)*20)"
    "+(((('Game Input'!C12)-8.9)/8.9)*10)"
    "+(((3.2-'Game Input'!C13)/3.2)*10)"
    "+(((4.20-'Game Input'!C17)/4.20)*5)"
    "+50))"
))

# ── Offense Scores ────────────────────────────

eng_label(39, "Home Offense Score")
eng_formula(40, (
    "=MAX(0,MIN(100,"
    "((('Game Input'!B21-100)/100)*40)"
    "+(((('Game Input'!B20)-0.315)/0.315)*35)"
    "+(((('Game Input'!B24)-8.5)/8.5)*15)"
    "+(((('Game Input'!B23)-0.315)/0.315)*10)"
    "+50))"
))

eng_label(41, "Away Offense Score")
eng_formula(42, (
    "=MAX(0,MIN(100,"
    "((('Game Input'!C21-100)/100)*40)"
    "+(((('Game Input'!C20)-0.315)/0.315)*35)"
    "+(((('Game Input'!C24)-8.5)/8.5)*15)"
    "+(((('Game Input'!C23)-0.315)/0.315)*10)"
    "+50))"
))

# ── Bullpen Scores ────────────────────────────

eng_label(43, "Home Bullpen Score")
eng_formula(44, (
    "=MAX(0,MIN(100,"
    "((4.30-'Game Input'!B27)/4.30)*60"
    "+((2.0-'Game Input'!B28)/2.0)*40"
    "+50))"
))

eng_label(45, "Away Bullpen Score")
eng_formula(46, (
    "=MAX(0,MIN(100,"
    "((4.30-'Game Input'!C27)/4.30)*60"
    "+((2.0-'Game Input'!C28)/2.0)*40"
    "+50))"
))

# ── Context Adjustments ───────────────────────

eng_label(47, "Park Factor Adj (Home)")
eng_formula(48, "=('Game Input'!B31-100)/100*0.3")

eng_label(49, "Wind Adj")
eng_formula(50, (
    '=IF(\'Game Input\'!B33="Out",\'Game Input\'!B32*0.02,'
    'IF(\'Game Input\'!B33="In",-\'Game Input\'!B32*0.015,0))'
))

eng_label(51, "Temp Adj")
eng_formula(52, (
    "=IF('Game Input'!B34<72,-(72-'Game Input'!B34)/10*0.005,0)"
))

eng_label(53, "Rest Adj (Home)")
eng_formula(54, (
    "=IF(('Game Input'!B35-'Game Input'!C36)>=2,0.03,"
    "IF(('Game Input'!C36-'Game Input'!B35)>=2,-0.03,0))"
))

# ── Expected Runs ─────────────────────────────

eng_label(55, "Expected Runs — Home")
eng_formula(56, (
    "=MAX(1.5,MIN(12,"
    "3.5"
    "+(B40/100)*3"
    "+(-B38/100)*1.5"
    "+(B44/100)*1"
    "+B48+B50+B52+B54"
    "))"
))

eng_label(57, "Expected Runs — Away")
eng_formula(58, (
    "=MAX(1.5,MIN(12,"
    "3.5"
    "+(B42/100)*3"
    "+(-B36/100)*1.5"
    "+(B46/100)*1"
    "+B48+B50+B52-B54"
    "))"
))

# ── Win Probability ───────────────────────────

eng_label(59, "Run Differential")
eng_formula(60, "=B56-B58")

eng_label(61, "Home Win Probability")
eng_formula(62, "=1/(1+EXP(-B60*0.8))")

eng_label(63, "Away Win Probability")
eng_formula(64, "=1-B62")

# ── Vegas Implied Probability ─────────────────

eng_label(65, "Home Implied Prob")
eng_formula(66, (
    "=IF('Game Input'!B39<0,"
    "ABS('Game Input'!B39)/(ABS('Game Input'!B39)+100),"
    "100/('Game Input'!B39+100))"
))

eng_label(67, "Away Implied Prob")
eng_formula(68, (
    "=IF('Game Input'!C39<0,"
    "ABS('Game Input'!C39)/(ABS('Game Input'!C39)+100),"
    "100/('Game Input'!C39+100))"
))

# ── Edge ──────────────────────────────────────

eng_label(69, "Home Edge %")
eng_formula(70, "=B62-B66")

eng_label(71, "Away Edge %")
eng_formula(72, "=B64-B68")

# ── Value Flag ────────────────────────────────

eng_label(73, "Home Value?")
eng_formula(74, '=IF(B70>0.05,"VALUE BET",IF(B70>0.02,"WATCH","NO VALUE"))')

eng_label(75, "Away Value?")
eng_formula(76, '=IF(B72>0.05,"VALUE BET",IF(B72>0.02,"WATCH","NO VALUE"))')

# ── Total Expected Runs ───────────────────────

eng_label(77, "Total Expected Runs")
eng_formula(78, "=B56+B58")

# ── Confidence Level ──────────────────────────

eng_label(79, "Confidence Level")
eng_formula(80, '=IF(MAX(B70,B72)>0.08,"High",IF(MAX(B70,B72)>0.04,"Medium","Low"))')

# ══════════════════════════════════════════════
# SHEET 3 — Results Dashboard
# ══════════════════════════════════════════════

ws_rd = wb.create_sheet("Results Dashboard")
ws_rd.tab_color = GREEN
ws_rd.freeze_panes = "A2"

set_col_width(ws_rd, "A", 25)
set_col_width(ws_rd, "B", 20)
set_col_width(ws_rd, "C", 10)
set_col_width(ws_rd, "D", 20)
set_col_width(ws_rd, "E", 20)
set_col_width(ws_rd, "F", 15)
set_col_width(ws_rd, "G", 15)

# Row 1 — header
apply_header(ws_rd, "A1", "MLB BETTING MODEL — RESULTS DASHBOARD", merge_to="G1", size=16)
ws_rd.row_dimensions[1].height = 30

# Row 2 — blank
ws_rd.row_dimensions[2].height = 8

# Row 3 — Team names
ws_rd["A3"].value = "Matchup:"
ws_rd["A3"].font = Font(bold=True, color="000000", size=12)
ws_rd["B3"].value = "='Game Input'!B4"
ws_rd["B3"].font = Font(bold=True, color=NAVY, size=12)
ws_rd["C3"].value = "VS"
ws_rd["C3"].font = Font(bold=True, color="000000", size=12)
ws_rd["C3"].alignment = center()
ws_rd["D3"].value = "='Game Input'!C4"
ws_rd["D3"].font = Font(bold=True, color=NAVY, size=12)

# Row 4 — blank
ws_rd.row_dimensions[4].height = 8

# Row 5 — MATCHUP sub-header
apply_section(ws_rd, "A5", "MATCHUP SUMMARY", merge_to="G5")

# Row 6 — MODEL WIN PROBABILITY header
apply_section(ws_rd, "A6", "MODEL WIN PROBABILITY", merge_to="G6")

# Row 7 — Win Prob values
for col, text in [("A7", ""), ("B7", "Home"), ("C7", ""), ("D7", "Away")]:
    ws_rd[col].value = text
    ws_rd[col].font = Font(bold=True, color="000000", size=11)
    ws_rd[col].alignment = center()
ws_rd["B7"].value = "=TEXT('Model Engine'!B62,\"0.0%\")"
ws_rd["B7"].font = Font(bold=True, color=GREEN, size=14)
ws_rd["B7"].alignment = center()
ws_rd["D7"].value = "=TEXT('Model Engine'!B64,\"0.0%\")"
ws_rd["D7"].font = Font(bold=True, color="C62828", size=14)
ws_rd["D7"].alignment = center()
ws_rd["A7"].value = "Home Win Prob"
ws_rd["C7"].value = "Away Win Prob"

# Row 8 — VEGAS IMPLIED PROBABILITY header
apply_section(ws_rd, "A8", "VEGAS IMPLIED PROBABILITY", merge_to="G8")

# Row 9 — Vegas implied values
ws_rd["A9"].value = "Home Implied"
ws_rd["A9"].font = Font(color="000000", size=11)
ws_rd["B9"].value = "=TEXT('Model Engine'!B66,\"0.0%\")"
ws_rd["B9"].font = Font(bold=True, color="000000", size=14)
ws_rd["B9"].alignment = center()
ws_rd["C9"].value = "Away Implied"
ws_rd["C9"].font = Font(color="000000", size=11)
ws_rd["D9"].value = "=TEXT('Model Engine'!B68,\"0.0%\")"
ws_rd["D9"].font = Font(bold=True, color="000000", size=14)
ws_rd["D9"].alignment = center()

# Row 10 — EDGE % header
apply_section(ws_rd, "A10", "EDGE %", merge_to="G10")

# Row 11 — Edge values with conditional formatting
ws_rd["A11"].value = "Home Edge"
ws_rd["A11"].font = Font(color="000000", size=11)
ws_rd["B11"].value = "='Model Engine'!B70"
ws_rd["B11"].number_format = "0.0%"
ws_rd["B11"].font = Font(bold=True, size=13)
ws_rd["B11"].alignment = center()
ws_rd["B11"].border = thin_border()

ws_rd["C11"].value = "Away Edge"
ws_rd["C11"].font = Font(color="000000", size=11)
ws_rd["D11"].value = "='Model Engine'!B72"
ws_rd["D11"].number_format = "0.0%"
ws_rd["D11"].font = Font(bold=True, size=13)
ws_rd["D11"].alignment = center()
ws_rd["D11"].border = thin_border()

# Conditional formatting for Edge cells
# Home Edge — B11
ws_rd.conditional_formatting.add("B11", FormulaRule(
    formula=["B11>0.05"], fill=fill(LIGHT_GREEN), font=Font(bold=True, color="1B5E20")
))
ws_rd.conditional_formatting.add("B11", FormulaRule(
    formula=["AND(B11>=0.02,B11<=0.05)"], fill=fill(LIGHT_ORANGE), font=Font(bold=True, color="E65100")
))
ws_rd.conditional_formatting.add("B11", FormulaRule(
    formula=["B11<0.02"], fill=fill(LIGHT_RED), font=Font(bold=True, color="B71C1C")
))

# Away Edge — D11
ws_rd.conditional_formatting.add("D11", FormulaRule(
    formula=["D11>0.05"], fill=fill(LIGHT_GREEN), font=Font(bold=True, color="1B5E20")
))
ws_rd.conditional_formatting.add("D11", FormulaRule(
    formula=["AND(D11>=0.02,D11<=0.05)"], fill=fill(LIGHT_ORANGE), font=Font(bold=True, color="E65100")
))
ws_rd.conditional_formatting.add("D11", FormulaRule(
    formula=["D11<0.02"], fill=fill(LIGHT_RED), font=Font(bold=True, color="B71C1C")
))

# Row 12 — blank
ws_rd.row_dimensions[12].height = 8

# Row 13 — RECOMMENDATION label
ws_rd["A13"].value = "RECOMMENDATION"
ws_rd["A13"].font = Font(bold=True, color="000000", size=12)
ws_rd["A13"].alignment = left()

# Row 14 — Recommendation formula (big, bold, conditional)
ws_rd.row_dimensions[14].height = 30
ws_rd.merge_cells("B14:E14")
ws_rd["B14"].value = (
    "=IF(AND('Model Engine'!B70>0.05,'Model Engine'!B70>'Model Engine'!B72),"
    "\"BET HOME\","
    "IF(AND('Model Engine'!B72>0.05,'Model Engine'!B72>'Model Engine'!B70),"
    "\"BET AWAY\",\"NO VALUE\"))"
)
ws_rd["B14"].font = Font(bold=True, size=14)
ws_rd["B14"].alignment = center()
ws_rd["B14"].border = thin_border()

# Conditional formatting for recommendation cell
ws_rd.conditional_formatting.add("B14", FormulaRule(
    formula=['B14="BET HOME"'], fill=fill(LIGHT_GREEN), font=Font(bold=True, color="1B5E20", size=14)
))
ws_rd.conditional_formatting.add("B14", FormulaRule(
    formula=['B14="BET AWAY"'], fill=fill(LIGHT_RED), font=Font(bold=True, color="B71C1C", size=14)
))
ws_rd.conditional_formatting.add("B14", FormulaRule(
    formula=['B14="NO VALUE"'], fill=fill(LIGHT_GRAY), font=Font(bold=True, color="424242", size=14)
))

# Row 15 — Confidence Level
ws_rd["A15"].value = "Confidence Level:"
ws_rd["A15"].font = Font(bold=True, color="000000", size=11)
ws_rd["B15"].value = "='Model Engine'!B80"
ws_rd["B15"].font = Font(bold=True, color=NAVY, size=12)
ws_rd["B15"].alignment = center()

# Row 16 — Expected Total Runs
ws_rd["A16"].value = "Expected Total Runs:"
ws_rd["A16"].font = Font(bold=True, color="000000", size=11)
ws_rd["B16"].value = "='Model Engine'!B78"
ws_rd["B16"].number_format = "0.00"
ws_rd["B16"].font = Font(bold=True, color=NAVY, size=12)
ws_rd["B16"].alignment = center()

# Row 17 — blank
ws_rd.row_dimensions[17].height = 8

# Row 18 — PITCHER COMPARISON header
apply_section(ws_rd, "A18", "PITCHER COMPARISON", merge_to="G18")

# Row 19 — Pitcher comparison column headers
for col, text in [("A19", "Stat"), ("B19", "Home Pitcher"), ("C19", "Away Pitcher")]:
    ws_rd[col].value = text
    ws_rd[col].fill = fill("E8EDF3")
    ws_rd[col].font = Font(bold=True, color="000000", size=11)
    ws_rd[col].alignment = center()
    ws_rd[col].border = thin_border()

pitcher_compare = [
    (20, "ERA",            "'Game Input'!B8",  "'Game Input'!C8"),
    (21, "FIP",            "'Game Input'!B9",  "'Game Input'!C9"),
    (22, "xFIP",           "'Game Input'!B10", "'Game Input'!C10"),
    (23, "SIERA",          "'Game Input'!B11", "'Game Input'!C11"),
    (24, "K/9",            "'Game Input'!B12", "'Game Input'!C12"),
    (25, "BB/9",           "'Game Input'!B13", "'Game Input'!C13"),
    (26, "Recent Form ERA", "'Game Input'!B17", "'Game Input'!C17"),
]
for row, stat, home_ref, away_ref in pitcher_compare:
    ws_rd[f"A{row}"].value = stat
    ws_rd[f"A{row}"].font = Font(color="000000", size=11)
    ws_rd[f"A{row}"].border = thin_border()
    ws_rd[f"A{row}"].fill = fill("F8FAFF")
    ws_rd[f"B{row}"].value = f"={home_ref}"
    ws_rd[f"B{row}"].font = Font(color="000000", size=11)
    ws_rd[f"B{row}"].alignment = center()
    ws_rd[f"B{row}"].border = thin_border()
    ws_rd[f"C{row}"].value = f"={away_ref}"
    ws_rd[f"C{row}"].font = Font(color="000000", size=11)
    ws_rd[f"C{row}"].alignment = center()
    ws_rd[f"C{row}"].border = thin_border()

# Row 27 — blank
ws_rd.row_dimensions[27].height = 8

# Row 28 — CONTEXT SUMMARY header
apply_section(ws_rd, "A28", "CONTEXT SUMMARY", merge_to="G28")

context_rows = [
    (29, "Park Factor",  "='Game Input'!B31"),
    (30, "Wind",         "=IF('Game Input'!B33=\"\",\"—\",'Game Input'!B33&\" \"&'Game Input'!B32&\" mph\")"),
    (31, "Temperature",  "='Game Input'!B34&\" °F\""),
    (32, "Home Rest",    "='Game Input'!B35&\" days\""),
    (33, "Away Rest",    "='Game Input'!C36&\" days\""),
]
for row, label, formula in context_rows:
    ws_rd[f"A{row}"].value = label
    ws_rd[f"A{row}"].font = Font(color="000000", size=11)
    ws_rd[f"A{row}"].fill = fill("F8FAFF")
    ws_rd[f"A{row}"].border = thin_border()
    ws_rd[f"B{row}"].value = formula
    ws_rd[f"B{row}"].font = Font(bold=True, color=NAVY, size=11)
    ws_rd[f"B{row}"].alignment = center()
    ws_rd[f"B{row}"].border = thin_border()

# ══════════════════════════════════════════════
# SHEET 4 — Bet Tracker
# ══════════════════════════════════════════════

ws_bt = wb.create_sheet("Bet Tracker")
ws_bt.tab_color = DARK_GOLD
ws_bt.freeze_panes = "A11"

set_col_width(ws_bt, "A", 12)
set_col_width(ws_bt, "B", 20)
set_col_width(ws_bt, "C", 20)
set_col_width(ws_bt, "D", 12)
set_col_width(ws_bt, "E", 15)
set_col_width(ws_bt, "F", 10)
set_col_width(ws_bt, "G", 10)
set_col_width(ws_bt, "H", 15)
set_col_width(ws_bt, "I", 18)
set_col_width(ws_bt, "J", 10)

# Row 1 — header
apply_header(ws_bt, "A1", "BET TRACKER", merge_to="J1", size=16)
ws_bt.row_dimensions[1].height = 30

# Row 2 — blank
ws_bt.row_dimensions[2].height = 8

# Row 3 — SUMMARY STATS
apply_section(ws_bt, "A3", "SUMMARY STATS", merge_to="E3")

# Rows 4–8 — summary formulas
summary = [
    (4, "Total Bets",        "=COUNTA(A11:A1000)",                               None),
    (5, "Win %",             "=IFERROR(COUNTIF(G11:G1000,\"W\")/COUNTA(G11:G1000),0)", "0.0%"),
    (6, "Total P&L",         "=SUM(H11:H1000)",                                  "$#,##0.00"),
    (7, "ROI %",             "=IFERROR(SUM(H11:H1000)/SUM(F11:F1000),0)",        "0.0%"),
    (8, "Current Bankroll",  "=1000+SUM(H11:H1000)",                             "$#,##0.00"),
]
for row, label, formula, num_fmt in summary:
    ws_bt[f"A{row}"].value = label
    ws_bt[f"A{row}"].font = Font(bold=True, color="000000", size=11)
    ws_bt[f"B{row}"].value = formula
    ws_bt[f"B{row}"].font = Font(bold=True, color=NAVY, size=12)
    ws_bt[f"B{row}"].alignment = center()
    ws_bt[f"B{row}"].border = thin_border()
    if num_fmt:
        ws_bt[f"B{row}"].number_format = num_fmt

# Row 9 — blank
ws_bt.row_dimensions[9].height = 8

# Row 10 — Column headers
tracker_headers = [
    "A10", "B10", "C10", "D10", "E10", "F10", "G10", "H10", "I10", "J10"
]
tracker_labels = [
    "Date", "Home Team", "Away Team", "Bet Side",
    "Moneyline Odds", "Stake ($)", "Result",
    "Profit/Loss", "Running Bankroll", "ROI %"
]
for cell_ref, label in zip(tracker_headers, tracker_labels):
    ws_bt[cell_ref].value = label
    ws_bt[cell_ref].fill = fill(NAVY)
    ws_bt[cell_ref].font = font(bold=True, color=WHITE, size=11)
    ws_bt[cell_ref].alignment = center()
    ws_bt[cell_ref].border = thin_border()
ws_bt.row_dimensions[10].height = 20

# Data validation for G column (Result): W / L / Push
dv_result = DataValidation(
    type="list",
    formula1='"W,L,Push"',
    showDropDown=False,
    allow_blank=True,
)
dv_result.sqref = "G11:G1000"
ws_bt.add_data_validation(dv_result)

# ── Example rows 11-13 ───────────────────────
example_rows = [
    (11, "5/1/2026", "New York Yankees",        "Boston Red Sox",          "HOME", -150, 100, "W"),
    (12, "5/2/2026", "Los Angeles Dodgers",      "San Diego Padres",        "AWAY", 130,   50, "L"),
    (13, "5/3/2026", "Atlanta Braves",           "Philadelphia Phillies",   "HOME", -120,  75, "W"),
]

for row, date, home, away, side, odds, stake, result in example_rows:
    ws_bt[f"A{row}"].value = date
    ws_bt[f"A{row}"].number_format = "MM/DD/YYYY"
    ws_bt[f"A{row}"].font = Font(color="000000", size=11)
    ws_bt[f"A{row}"].border = thin_border()

    ws_bt[f"B{row}"].value = home
    ws_bt[f"B{row}"].font = Font(color="000000", size=11)
    ws_bt[f"B{row}"].border = thin_border()

    ws_bt[f"C{row}"].value = away
    ws_bt[f"C{row}"].font = Font(color="000000", size=11)
    ws_bt[f"C{row}"].border = thin_border()

    ws_bt[f"D{row}"].value = side
    ws_bt[f"D{row}"].font = Font(color="000000", size=11)
    ws_bt[f"D{row}"].alignment = center()
    ws_bt[f"D{row}"].border = thin_border()

    ws_bt[f"E{row}"].value = odds
    ws_bt[f"E{row}"].font = Font(color="000000", size=11)
    ws_bt[f"E{row}"].alignment = center()
    ws_bt[f"E{row}"].border = thin_border()

    ws_bt[f"F{row}"].value = stake
    ws_bt[f"F{row}"].number_format = "$#,##0.00"
    ws_bt[f"F{row}"].font = Font(color="000000", size=11)
    ws_bt[f"F{row}"].alignment = center()
    ws_bt[f"F{row}"].border = thin_border()

    ws_bt[f"G{row}"].value = result
    ws_bt[f"G{row}"].font = Font(bold=True, color="000000", size=11)
    ws_bt[f"G{row}"].alignment = center()
    ws_bt[f"G{row}"].border = thin_border()

    # H — Profit/Loss formula
    ws_bt[f"H{row}"].value = (
        f"=IF(G{row}=\"W\","
        f"IF(E{row}<0,F{row}/ABS(E{row})*100,F{row}*E{row}/100),"
        f"IF(G{row}=\"L\",-F{row},\"\"))"
    )
    ws_bt[f"H{row}"].number_format = "$#,##0.00"
    ws_bt[f"H{row}"].font = Font(color="000000", size=11)
    ws_bt[f"H{row}"].alignment = center()
    ws_bt[f"H{row}"].border = thin_border()

    # I — Running Bankroll
    ws_bt[f"I{row}"].value = f"=1000+SUM(H$11:H{row})"
    ws_bt[f"I{row}"].number_format = "$#,##0.00"
    ws_bt[f"I{row}"].font = Font(color="000000", size=11)
    ws_bt[f"I{row}"].alignment = center()
    ws_bt[f"I{row}"].border = thin_border()

    # J — ROI %
    ws_bt[f"J{row}"].value = f"=IFERROR(SUM(H$11:H{row})/SUM(F$11:F{row}),\"\")"
    ws_bt[f"J{row}"].number_format = "0.0%"
    ws_bt[f"J{row}"].font = Font(color="000000", size=11)
    ws_bt[f"J{row}"].alignment = center()
    ws_bt[f"J{row}"].border = thin_border()

# ── Rows 14–200: pre-format formulas ─────────
for row in range(14, 201):
    # Profit/Loss
    ws_bt[f"H{row}"].value = (
        f"=IF(G{row}=\"W\","
        f"IF(E{row}<0,F{row}/ABS(E{row})*100,F{row}*E{row}/100),"
        f"IF(G{row}=\"L\",-F{row},\"\"))"
    )
    ws_bt[f"H{row}"].number_format = "$#,##0.00"
    ws_bt[f"H{row}"].font = Font(color="000000", size=11)
    ws_bt[f"H{row}"].alignment = center()
    ws_bt[f"H{row}"].border = thin_border()
    ws_bt[f"H{row}"].fill = fill(LIGHT_GRAY)

    # Running Bankroll
    ws_bt[f"I{row}"].value = f"=IF(COUNTA(A$11:A{row})=0,\"\",1000+SUM(H$11:H{row}))"
    ws_bt[f"I{row}"].number_format = "$#,##0.00"
    ws_bt[f"I{row}"].font = Font(color="000000", size=11)
    ws_bt[f"I{row}"].alignment = center()
    ws_bt[f"I{row}"].border = thin_border()
    ws_bt[f"I{row}"].fill = fill(LIGHT_GRAY)

    # ROI %
    ws_bt[f"J{row}"].value = f"=IFERROR(SUM(H$11:H{row})/SUM(F$11:F{row}),\"\")"
    ws_bt[f"J{row}"].number_format = "0.0%"
    ws_bt[f"J{row}"].font = Font(color="000000", size=11)
    ws_bt[f"J{row}"].alignment = center()
    ws_bt[f"J{row}"].border = thin_border()
    ws_bt[f"J{row}"].fill = fill(LIGHT_GRAY)

# ══════════════════════════════════════════════
# SHEET 5 — Reference & Glossary
# ══════════════════════════════════════════════

ws_ref = wb.create_sheet("Reference & Glossary")
ws_ref.tab_color = BLUE
ws_ref.freeze_panes = "A2"

set_col_width(ws_ref, "A", 18)
set_col_width(ws_ref, "B", 60)
set_col_width(ws_ref, "C", 16)
set_col_width(ws_ref, "D", 30)

# Row 1 — header
apply_header(ws_ref, "A1", "REFERENCE & GLOSSARY", merge_to="D1", size=16)
ws_ref.row_dimensions[1].height = 30

# Row 2 — blank
ws_ref.row_dimensions[2].height = 8

# Row 3 — section label
apply_section(ws_ref, "A3", "METRIC DEFINITIONS", merge_to="D3")

# Row 4 — column headers
for col, text in [("A4", "Metric"), ("B4", "Definition"), ("C4", "League Average"), ("D4", "Where to Find It")]:
    ws_ref[col].value = text
    ws_ref[col].fill = fill("E8EDF3")
    ws_ref[col].font = Font(bold=True, color="000000", size=11)
    ws_ref[col].alignment = center()
    ws_ref[col].border = thin_border()

glossary = [
    (
        "FIP",
        "Fielding Independent Pitching. Measures pitcher performance based only on strikeouts, walks, and home runs — removes defense from the equation. Lower is better.",
        "4.20",
        "FanGraphs Leaderboards"
    ),
    (
        "xFIP",
        "Expected FIP. Like FIP but normalizes HR/FB rate to league average. Better predictor of future performance than ERA. Lower is better.",
        "4.15",
        "FanGraphs Leaderboards"
    ),
    (
        "SIERA",
        "Skill-Interactive ERA. Most advanced ERA estimator. Accounts for batted ball quality. Lower is better.",
        "4.10",
        "FanGraphs Leaderboards"
    ),
    (
        "K/9",
        "Strikeouts per 9 innings. Higher means more swing-and-miss ability. Reduces opponent's ability to put the ball in play.",
        "8.9",
        "FanGraphs / Baseball Reference"
    ),
    (
        "BB/9",
        "Walks per 9 innings. Lower is better. High BB/9 indicates control issues that inflate scoring.",
        "3.2",
        "FanGraphs / Baseball Reference"
    ),
    (
        "wOBA",
        "Weighted On-Base Average. Weights each way of reaching base by its actual run value. Better than OBP or AVG alone.",
        ".315",
        "FanGraphs Leaderboards"
    ),
    (
        "wRC+",
        "Weighted Runs Created Plus. Park- and league-adjusted offensive metric. 100 = league average. 120 = 20% above average.",
        "100",
        "FanGraphs Leaderboards"
    ),
    (
        "BABIP",
        "Batting Average on Balls In Play. Measures luck/defense for pitchers, luck/skill for hitters. Extreme values tend to regress.",
        ".300",
        "FanGraphs / Baseball Reference"
    ),
    (
        "Barrel Rate",
        "% of batted balls hit with optimal exit velocity and launch angle. Strongly correlates with future offensive production.",
        "8.5%",
        "Baseball Savant"
    ),
    (
        "Park Factor",
        "Adjusts for how a ballpark affects run scoring. 100 = neutral. Above 100 = hitter-friendly, below 100 = pitcher-friendly.",
        "100",
        "Baseball Reference Park Factors"
    ),
    (
        "Edge %",
        "Model Win Probability minus Vegas Implied Probability. Positive edge means the model thinks you have an advantage vs the line.",
        "N/A",
        "Calculated by this model"
    ),
    (
        "Win Probability",
        "Model's predicted probability of winning based on all input factors. Output of the logistic regression calculation.",
        "50%",
        "Calculated by this model"
    ),
    (
        "Implied Probability",
        "The win probability implied by the Vegas moneyline. -150 implies 60% chance of winning.",
        "N/A",
        "Calculated from moneyline"
    ),
]

for i, (metric, definition, avg, source) in enumerate(glossary, start=5):
    row = i
    bg = "F8FAFF" if i % 2 == 0 else WHITE
    for col in range(1, 5):
        ws_ref.cell(row=row, column=col).fill = fill(bg)
        ws_ref.cell(row=row, column=col).border = thin_border()
        ws_ref.cell(row=row, column=col).font = Font(color="000000", size=10)

    ws_ref.cell(row=row, column=1, value=metric)
    ws_ref.cell(row=row, column=1).font = Font(bold=True, color=NAVY, size=10)
    ws_ref.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws_ref.cell(row=row, column=2, value=definition)
    ws_ref.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws_ref.cell(row=row, column=3, value=avg)
    ws_ref.cell(row=row, column=3).alignment = center()

    ws_ref.cell(row=row, column=4, value=source)
    ws_ref.cell(row=row, column=4).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws_ref.row_dimensions[row].height = 40

# ── Section 2: How to Use ─────────────────────
how_to_row_start = 5 + len(glossary) + 2   # a gap row
gap_row = 5 + len(glossary)
ws_ref.row_dimensions[gap_row].height = 10

how_to_header_row = gap_row + 1
apply_section(ws_ref, f"A{how_to_header_row}", "HOW TO USE THE MODEL", merge_to=f"D{how_to_header_row}")

steps = [
    "Step 1: Go to the 'Game Input' sheet.",
    "Step 2: Select Home and Away teams from the dropdown menus (Park Factor auto-populates).",
    "Step 3: Enter pitcher stats from FanGraphs.com → Pitchers → Leaderboards.",
    "Step 4: Enter team offensive stats from FanGraphs.com → Teams → Offensive stats.",
    "Step 5: Enter bullpen stats from FanGraphs.com → Teams → Relief pitchers (filter last 14 days).",
    "Step 6: Enter context factors (wind, temp from weather.com; rest days from team schedule).",
    "Step 7: Enter Vegas moneyline from your preferred sportsbook.",
    "Step 8: Navigate to 'Results Dashboard' to see the model output.",
    "Step 9: Only bet when Edge % > 5% and confidence is Medium or High.",
    "Step 10: Log every bet in the 'Bet Tracker' sheet to monitor ROI.",
]

for step_i, step_text in enumerate(steps, start=1):
    r = how_to_header_row + step_i
    ws_ref.merge_cells(f"A{r}:D{r}")
    ws_ref[f"A{r}"].value = step_text
    ws_ref[f"A{r}"].font = Font(color="000000", size=11)
    ws_ref[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_ref[f"A{r}"].border = thin_border()
    if step_i % 2 == 0:
        ws_ref[f"A{r}"].fill = fill("F8FAFF")
    ws_ref.row_dimensions[r].height = 20

# ══════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════

output_path = "/Users/jameson/Desktop/MLB-Betting-model/MLB_Betting_Model.xlsx"
wb.save(output_path)

file_size = os.path.getsize(output_path)
print(f"SUCCESS: MLB_Betting_Model.xlsx saved to {output_path}")
print(f"File size: {file_size:,} bytes ({file_size / 1024:.1f} KB)")
print("Sheets created:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
